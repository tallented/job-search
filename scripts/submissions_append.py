#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DATE_FORMAT = "mm/dd/yy;@"


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    text = value.strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value!r}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Append a row to the archived Submissions sheet in 2026 Job Search.xlsx. "
            "Use this only for legacy workbook maintenance when the workbook is closed in Excel."
        )
    )
    parser.add_argument("--workbook", default="/Users/chris/2026 Resumes/bak/2026 Job Search.xlsx")
    parser.add_argument("--sheet", default="Submissions")
    parser.add_argument("--company", required=True)
    parser.add_argument("--status", default="CREATED")
    parser.add_argument("--created")
    parser.add_argument("--submitted")
    parser.add_argument("--title", required=True)
    parser.add_argument("--type")
    parser.add_argument("--location")
    parser.add_argument("--salary")
    parser.add_argument("--next-steps")
    parser.add_argument("--listing-source")
    parser.add_argument("--ref-number")
    parser.add_argument("--link")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    ws = wb[args.sheet]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(value): idx for idx, value in enumerate(headers, start=1) if value}

    required_headers = [
        "Company",
        "Status",
        "Created",
        "Submitted",
        "Title",
        "Type",
        "Location",
        "Salary",
        "Next Steps",
        "Listing Source",
        "Ref Number",
        "Link",
    ]
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        raise KeyError(f"Missing expected headers: {missing}")

    new_row = ws.max_row + 1
    source_row = ws.max_row

    # Copy formatting from the previous row so borders/fills/alignment remain consistent.
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=new_row, column=col)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)

    created = parse_date(args.created) or date.today()
    submitted = parse_date(args.submitted)

    values = {
        "Company": args.company,
        "Status": args.status,
        "Created": created,
        "Submitted": submitted,
        "Title": args.title,
        "Type": args.type,
        "Location": args.location,
        "Salary": args.salary,
        "Next Steps": args.next_steps,
        "Listing Source": args.listing_source,
        "Ref Number": args.ref_number,
        "Link": args.link,
    }

    for header, value in values.items():
        cell = ws.cell(row=new_row, column=header_map[header])
        cell.value = value
        if header in {"Created", "Submitted"}:
            cell.number_format = DATE_FORMAT

    wb.save(workbook_path)
    print(f"Appended row {new_row} to {workbook_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
